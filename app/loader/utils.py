import numpy as np
from matplotlib.dates import date2num, num2date
from openpyxl import load_workbook
from units.models import Unit, UnitCreator, UnitAction
from aircarts.models import Aircart, AircartCompany, AircartFlightRecord
from django.db.models import F, Sum
from django.db import transaction


#Извлекаем контент из xlsx
def preprocess_xlsx(filename):
    wb = load_workbook(filename=filename)
    sheet = wb.active
    main = {}
    # TODO review this structure (what is the purpose)
    ll = {'FH': ['name', 'source', 'statistic'],
        'Removals': ['name', 'source', 'statistic'],
        'Failures': ['name', 'source', 'statistic'],
        'Induced': ['name', 'source', 'statistic']}
    for key in wb.get_sheet_names():
        main_dict = []
        current_wb = wb[key]
        for i in np.arange(current_wb.min_row + 1, current_wb.max_row):
            buff_dict = {}
            stat_dict = np.matrix([[date2num(current_wb.cell(row=current_wb.min_row, column=j).value),
                                   current_wb.cell(row=i, column=j).value if current_wb.cell(row=i, column=j).value != None else 0
                                    ] for j in np.arange(current_wb.min_column + 2, current_wb.max_column)])
            buff_dict[ll[key][0]] = current_wb.cell(row=i, column=1).value
            buff_dict[ll[key][1]] = current_wb.cell(row=i, column=2).value
            buff_dict[ll[key][2]] = np.array(stat_dict)
            main_dict.append(buff_dict)
            main[key] = main_dict
    return main

#Сохраняем статистику из словаря в бд
def store_to_db(data):
    #Идем по "Листам"
    for leaf_name, leaf_content in data.items():
        if leaf_name == 'FH':
            #Сохраняем FH
            for leaf_row in leaf_content:
                #Проверяем на создание компанию
                source = AircartCompany.objects.get_or_create(name=leaf_row['source'])
                #Проверяем на создание самолет
                obj = Aircart.objects.get_or_create(company=source[0], number=leaf_row['name'])
                #Идем по статистике
                for event in leaf_row['statistic']:
                    #Проверка на пустоту статистики
                    if event[1]:
                        #Проверка на наличие в базе
                        if AircartFlightRecord.objects.filter(aircart=obj[0], date=num2date(event[0])).exists():
                            AircartFlightRecord.objects.update_or_create(aircart=obj[0], date=num2date(event[0]), defaults={'count': F('count') + event[1]})
                        else:
                            AircartFlightRecord.objects.create(aircart=obj[0], date=num2date(event[0]), count=event[1])
        else:
            #Заполняем статистику для блоков
            for leaf_row in leaf_content:
                if leaf_name == 'Removals':
                    mark = 0
                elif leaf_name == 'Failures':
                    mark = 1   
                elif leaf_name == 'Induced':
                    mark = 2
                #Снова проверяем на объект/источник
                source = UnitCreator.objects.get_or_create(name=leaf_row['source'])
                obj = Unit.objects.get_or_create(manufacturer=source[0], number=leaf_row['name'])
                for event in leaf_row['statistic']:
                    #Вставляем failtures/removals столько раз сколько было описано в excel
                    for _ in np.arange(event[1]):
                        UnitAction.objects.create(unit=obj[0], date=num2date(event[0]), action_type=mark)