import numpy as np
from matplotlib.dates import date2num, num2date
from openpyxl import load_workbook
from units.models import Unit, UnitCreator, UnitAction
from aircarts.models import Plane, PlaneCompany, PlaneFlightHours
from django.db.models import F, Sum


def is_source_exist(source, Model):
    return Model.objects.filter(source=source).exists()

def is_name_exist(name, Model):
    return Model.objects.filter(name=name).exists()

def preprocess_xlsx(filename):
    wb = load_workbook(filename=filename)
    sheet = wb.active
    main = {}
    # TODO review this structure (what is the purpose)
    ll = {'FH': ['name', 'source', 'statistic'],
        'Removals': ['name', 'source', 'statistic'],
        'Failures': ['name', 'source', 'statistic'],
        'Induced': ['name', 'source', 'statistic']}
    for key in wb.get_sheet_names():
        main_dict = []
        for i in np.arange(wb.get_sheet_by_name(key).min_row + 1, wb.get_sheet_by_name(key).max_row):
            buff_dict = {}
            stat_dict = np.matrix([[date2num(wb.get_sheet_by_name(key).cell(row=wb.get_sheet_by_name(key).min_row, column=j).value),
                                    wb.get_sheet_by_name(key).cell(row=i, column=j).value if wb.get_sheet_by_name(key).cell(row=i, column=j).value != None else 0
                                    ] for j in np.arange(wb.get_sheet_by_name(key).min_column + 2, wb.get_sheet_by_name(key).max_column)])
            # stat_dict = []
            # for j in range(wb.get_sheet_by_name(key).min_column + 2, wb.get_sheet_by_name(key).max_column):
            #     stat_dict.append(np.array([wb.get_sheet_by_name(key).cell(row=wb.get_sheet_by_name(key).min_row, column=j).value,
            #                       wb.get_sheet_by_name(key).cell(row=i, column=j).value if wb.get_sheet_by_name(key).cell(row=i, column=j).value != None else 0]))
            buff_dict[ll[key][0]] = wb.get_sheet_by_name(key).cell(row=i, column=1).value
            buff_dict[ll[key][1]] = wb.get_sheet_by_name(key).cell(row=i, column=2).value
            buff_dict[ll[key][2]] = np.array(stat_dict)
            main_dict.append(buff_dict)
            main[key] = main_dict
    return main

def store_to_db(data):
    for leaf_name, leaf_content in data.items():
        if leaf_name == 'FH':
            for leaf_row in leaf_content:
                source = PlaneCompany.objects.get_or_create(name=leaf_row['source'])
                obj = Plane.objects.get_or_create(company_id=source[0], board_number=leaf_row['name'])
                for event in leaf_row['statistic']:
                    if event[1]!=0:
                        if PlaneFlightHours.objects.filter(plane_id=obj[0], date=num2date(event[0])).exists():
                            PlaneFlightHours.objects.update_or_create(plane_id=obj[0], date=num2date(event[0]), defaults={'count': F('count') + event[1]})
                        else:
                            PlaneFlightHours.objects.create(plane_id=obj[0], date=num2date(event[0]), count=event[1])
        else:
            for leaf_row in leaf_content:
                if leaf_name == 'Removals':
                    mark = 0
                elif leaf_name == 'Failures':
                    mark = 1   
                elif leaf_name == 'Induced':
                    mark = 2
                source = UnitCreator.objects.get_or_create(name=leaf_row['source'])
                obj = Unit.objects.get_or_create(manufacturer_id=source[0], unit_number=leaf_row['name'])
                for event in leaf_row['statistic']:
                    for _ in np.arange(event[1]):
                        UnitAction.objects.create(unit_id=obj[0], date=num2date(event[0]), action_type=mark)