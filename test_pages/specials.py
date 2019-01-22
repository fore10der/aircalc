from openpyxl import load_workbook
from datetime import date, datetime
from django.http import HttpResponse
from django.shortcuts import render
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.dates import date2num, num2date
import io
import base64
from xhtml2pdf import pisa
from django.template.loader import get_template
from units.models import Unit, UnitCreator, UnitAction
from aircarts.models import Plane, PlaneCompany, PlaneFlightHours
from django.db.models import F

def is_source_exist(source, Model):
    return Model.objects.filter(source=source).exists()

def is_name_exist(name, Model):
    return Model.objects.filter(name=name).exists()

def build_bars(data):
    context = []
    for table_key, table_values in data.items():
        for item_values in table_values:
            ax = plt.subplot(111)
            ax.bar(num2date(item_values['statistic'][:,0]), item_values['statistic'][:,1], width=30)
            ax.xaxis_date()
            plt.title(item_values["name"])
            buf = io.BytesIO()
            plt.savefig(buf, format='png', dpi=100)
            image_base64 = base64.b64encode(buf.getvalue()).decode('utf-8').replace('\n', '')
            context.append(image_base64)
            buf.close()
            print(context)
            break
        break
    return context

def build_pdf(true_context):
    template_path = 'pdf.html'
    context = true_context
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="report.pdf"'
    # find the template and render it.
    template = get_template(template_path)
    html = template.render(context)

    # create a pdf
    pisaStatus = pisa.CreatePDF(
       html, dest=response)
    # if error then show some funy view
    print(pisaStatus.err)
    if pisaStatus.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response

def preprocess_xlsx(filename):

    wb = load_workbook(filename=filename)

    sheet = wb.active

    main = {}
    # TODO review this structure (what is the purpose)
    ll = {'FH': ['name', 'source', 'statistic'],
        'Removals': ['name', 'source', 'statistic'],
        'Failures': ['name', 'source', 'statistic'],
        'Induced': ['name', 'source', 'statistic']}
    for key in wb.get_sheet_names():
        main_dict = []
        for i in np.arange(wb.get_sheet_by_name(key).min_row + 1, wb.get_sheet_by_name(key).max_row):
            buff_dict = {}
            stat_dict = np.matrix([[date2num(wb.get_sheet_by_name(key).cell(row=wb.get_sheet_by_name(key).min_row, column=j).value),
                                    wb.get_sheet_by_name(key).cell(row=i, column=j).value if wb.get_sheet_by_name(key).cell(row=i, column=j).value != None else 0
                                    ] for j in np.arange(wb.get_sheet_by_name(key).min_column + 2, wb.get_sheet_by_name(key).max_column)])
            # stat_dict = []
            # for j in range(wb.get_sheet_by_name(key).min_column + 2, wb.get_sheet_by_name(key).max_column):
            #     stat_dict.append(np.array([wb.get_sheet_by_name(key).cell(row=wb.get_sheet_by_name(key).min_row, column=j).value,
            #                       wb.get_sheet_by_name(key).cell(row=i, column=j).value if wb.get_sheet_by_name(key).cell(row=i, column=j).value != None else 0]))
            buff_dict[ll[key][0]] = wb.get_sheet_by_name(key).cell(row=i, column=1).value
            buff_dict[ll[key][1]] = wb.get_sheet_by_name(key).cell(row=i, column=2).value
            buff_dict[ll[key][2]] = np.array(stat_dict)
            main_dict.append(buff_dict)
            main[key] = main_dict
    
    return main

def store_to_db(data):
    for leaf_name, leaf_content in data.items():
        if leaf_name == 'FH':
            for leaf_row in leaf_content:
                source = PlaneCompany.objects.get_or_create(name=leaf_row['source'])
                obj = Plane.objects.get_or_create(company_id=source[0], board_number=leaf_row['name'])
                for event in leaf_row['statistic']:
                    if event[1]!=0:
                        if PlaneFlightHours.objects.filter(plane_id=obj[0], date=num2date(event[0])).exists():
                            PlaneFlightHours.objects.update_or_create(plane_id=obj[0], date=num2date(event[0]), defaults={'count': F('count') + event[1]})
                        else:
                            PlaneFlightHours.objects.create(plane_id=obj[0], date=num2date(event[0]), count=event[1])
        elif leaf_name == 'Failures' or leaf_name == 'Induced':
            for leaf_row in leaf_content:
                source = UnitCreator.objects.get_or_create(name=leaf_row['source'])
                obj = Unit.objects.get_or_create(manufacturer_id=source[0], unit_number=leaf_row['name'])
                for event in leaf_row['statistic']:
                    mark = 0 if leaf_name == 'Induced' else 1
                    for _ in np.arange(event[1]):
                        UnitAction.objects.create(unit_id=obj[0], date=num2date(event[0]), action_type=mark)